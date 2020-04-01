import pandas as pd
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
import os.path
from io import BytesIO

# Exports the data from either descriptive or basic statistic's datatable
# Saves to a temporary excel file
# Returns a str_io file for download
def exportToFile():
    wb = Workbook()

    ws = wb.active
    ws.title = "Dataset"

    #Adds the dataframe values to the sheet
    #for r in dataframe_to_rows(df, index=True, header=True):
        #ws.append(r)

    #highlights the top row
    #for cell in ws['A'] + ws[1]:
     #   cell.style = 'Pandas'

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        str_io = BytesIO(tmp.read())
    return str_io
    